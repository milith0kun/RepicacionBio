import io
from openpyxl import load_workbook

from ovo import db, Design, Project, Pool, Round, DescriptorValue, DesignSpec, DesignChain
from ovo.core.database.descriptors_proteinqc import PROTEINQC_MAIN_DESCRIPTORS
from ovo.core.logic.descriptor_logic import export_proteinqc_excel


def create_test_data():
    # Create minimal test data for export, matching import/export test conventions
    project = Project(id="test_proteinqc_excel_project", name="Test Excel Project", author="test_user", public=False)
    db.save(project)

    round_ = Round(id="test_proteinqc_excel_round", project_id=project.id, name="Test Round", author="test_user")
    db.save(round_)

    pool = Pool(id="tep", round_id=round_.id, name="Test Pool", author="test_user")
    db.save(pool)

    spec = DesignSpec(chains=[DesignChain(type="protein", chain_ids=["A"], sequence="ACDEFGHIK")])

    design = Design(id=f"ovo_{pool.id}_test_proteinqc_excel_design", pool_id=pool.id, accepted=True, spec=spec)
    db.save(design)

    # Add a DescriptorValue for the design
    descriptor_value = DescriptorValue(
        design_id=design.id,
        descriptor_key="proteinqc|seq_composition|length",
        descriptor_job_id="test_job_id",
        chains="A",
        value="10",
    )
    db.save(descriptor_value)

    return design


def test_export_proteinqc_excel():
    """
    Exports the proteinqc excel, reads it back and checks the descriptor value of the design.
    """
    design = create_test_data()
    excel_data = export_proteinqc_excel([design.id])
    wb = load_workbook(excel_data)
    ws = wb.active
    headers = [cell.value for cell in ws[2]]
    assert "Sequence length" in headers, "Descriptor column missing in Excel export"
    idx = headers.index("Sequence length")
    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == design.id:
            found = True
            assert str(row[idx]) == "10", f"Expected descriptor value '10', got {row[idx]}"
            break
    assert found, "Design row not found in exported Excel"
